"""XLSX (Excel) converter with multi-sheet support."""

from pathlib import Path
from typing import Optional, List, Dict
from loguru import logger

from converters.base_converter import DocumentContent


class XlsxConverter:
    """Convert XLSX files (Excel) to markdown with multi-sheet support."""
    
    def __init__(self, config: dict):
        """Initialize XLSX converter.
        
        Args:
            config: Configuration dictionary
        """
        self.config = config
        self.max_rows_per_sheet = config.get('xlsx', {}).get('max_rows', 1000)
        self.include_formulas = config.get('xlsx', {}).get('include_formulas', False)
    
    def convert(self, file_path: Path) -> DocumentContent:
        """Convert XLSX file to DocumentContent.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            DocumentContent object with all sheets
        """
        logger.info(f"Converting XLSX: {file_path.name}")
        
        try:
            # Try to import openpyxl
            try:
                import openpyxl
                import pandas as pd
            except ImportError:
                logger.warning("openpyxl or pandas not installed")
                content = DocumentContent()
                content.title = file_path.stem
                content.text = f"# {file_path.stem}\n\n**Error:** openpyxl library not installed.\n\nInstall with: pip install openpyxl"
                return content
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Get sheet information
            sheet_names = wb.sheetnames
            sheet_count = len(sheet_names)
            
            logger.info(f"XLSX has {sheet_count} sheets: {sheet_names}")
            
            # Build markdown content
            content_parts = []
            
            # Title
            content_parts.append(f"# {file_path.stem}\n")
            content_parts.append(f"**Excel Workbook** with {sheet_count} sheet(s)\n")
            
            # Table of contents
            if sheet_count > 1:
                content_parts.append("## Sheets\n")
                for sheet_name in sheet_names:
                    content_parts.append(f"- [{sheet_name}](#{sheet_name.lower().replace(' ', '-')})")
                content_parts.append("")
            
            # Process each sheet
            total_rows = 0
            total_cols = 0
            
            for sheet_name in sheet_names:
                logger.info(f"Processing sheet: {sheet_name}")
                
                sheet = wb[sheet_name]
                
                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                total_rows += max_row
                total_cols = max(total_cols, max_col)
                
                # Add sheet header
                content_parts.append(f"\n## {sheet_name}\n")
                content_parts.append(f"**Dimensions:** {max_row} rows × {max_col} columns\n")
                
                # Check if sheet is too large
                if max_row > self.max_rows_per_sheet:
                    content_parts.append(f"*Note: Sheet has {max_row} rows. Showing first {self.max_rows_per_sheet} rows.*\n")
                    max_row = self.max_rows_per_sheet
                
                # Convert to pandas DataFrame for easy markdown table generation
                try:
                    # Read sheet data
                    data = []
                    headers = []
                    
                    # Get headers from first row
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(1, col).value
                        headers.append(str(cell_value) if cell_value is not None else f"Column{col}")
                    
                    # Get data rows
                    for row in range(2, min(max_row + 1, self.max_rows_per_sheet + 1)):
                        row_data = []
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row, col)
                            value = cell.value
                            
                            # Handle different cell types
                            if value is None:
                                row_data.append('')
                            elif isinstance(value, (int, float)):
                                row_data.append(value)
                            else:
                                row_data.append(str(value))
                        
                        data.append(row_data)
                    
                    # Create DataFrame
                    if data:
                        df = pd.DataFrame(data, columns=headers)
                        
                        # Convert to markdown table
                        markdown_table = df.to_markdown(index=False)
                        content_parts.append(markdown_table)
                        
                        # Add statistics for numeric columns
                        numeric_cols = df.select_dtypes(include=['number']).columns
                        if len(numeric_cols) > 0 and len(df) > 1:
                            content_parts.append(f"\n### Statistics\n")
                            stats_df = df[numeric_cols].describe()
                            content_parts.append(stats_df.to_markdown())
                    else:
                        content_parts.append("*(Empty sheet)*")
                
                except Exception as e:
                    logger.warning(f"Failed to convert sheet {sheet_name}: {e}")
                    content_parts.append(f"*(Error converting sheet: {str(e)})*")
                
                content_parts.append("")
            
            wb.close()
            
            full_text = "\n".join(content_parts)
            
            # Create DocumentContent
            content = DocumentContent()
            content.title = file_path.stem
            content.text = full_text
            content.metadata = {
                'sheet_count': sheet_count,
                'sheet_names': sheet_names,
                'total_rows': total_rows,
                'max_columns': total_cols,
                'file_size': file_path.stat().st_size
            }
            
            logger.success(f"XLSX converted: {sheet_count} sheets, {total_rows} total rows")
            return content
            
        except Exception as e:
            logger.error(f"Failed to convert XLSX {file_path.name}: {e}")
            content = DocumentContent()
            content.title = file_path.stem
            content.text = f"# {file_path.stem}\n\nError converting XLSX: {str(e)}"
            content.metadata = {'error': str(e)}
            return content

